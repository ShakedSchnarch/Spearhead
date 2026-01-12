from pathlib import Path

from spearhead.data.adapters import (
    PlatoonLoadoutAdapter,
    BattalionSummaryAdapter,
    FormResponsesAdapter,
)
from openpyxl import load_workbook


BASE = Path(__file__).resolve().parents[1]


def test_platoon_loadout_zivud_and_ammo():
    path = BASE / "docs/Files/דוחות פלוגת כפיר.xlsx"
    records = PlatoonLoadoutAdapter.load(path)
    assert records, "No records parsed from platoon loadout file"

    zivud_match = [
        r for r in records
        if r.section == "zivud" and r.item == "חבל פריסה" and r.column == "צ'636"
    ]
    assert zivud_match, "Expected zivud record for צ'636 חבל פריסה"
    assert zivud_match[0].value == 1.0

    ammo_match = [
        r for r in records
        if r.section == "ammo" and r.item == "חלול"
    ]
    assert ammo_match, "Expected ammo record for חלול"
    # Values should be numeric; accept float/int
    assert any(float(m.value) > 0 for m in ammo_match)


def test_battalion_summary_parses_gaps():
    path = BASE / "docs/Files/מסמך דוחות גדודי.xlsx"
    records = BattalionSummaryAdapter.load(path)
    assert records, "No records parsed from battalion summary"

    gap = [
        r for r in records
        if r.section == "summary_zivud" and r.item == "חבל פריסה" and "סופה" in r.column
    ]
    assert gap, "Expected summary gap for חבל פריסה (סופה)"
    assert "בלאי" in str(gap[0].value)


def test_form_responses_status_present():
    path = BASE / "docs/Files/טופס דוחות סמפ כפיר. (תגובות).xlsx"
    responses = FormResponsesAdapter.load(path)
    assert responses, "No responses parsed from form export"
    first = responses[0]
    assert first.fields.get("דוח זיווד [חבל פריסה]") == "קיים"


def test_form_responses_aliases_and_unmapped(tmp_path):
    """
    Headers with extra spacing/punctuation should still resolve via config, while unknown columns are tracked.
    """
    src = BASE / "docs/Files/טופס דוחות סמפ כפיר. (תגובות).xlsx"
    wb = load_workbook(src)
    ws = wb.active

    def replace_header(original: str, new: str):
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value == original:
                cell.value = new
                return

    replace_header("צ טנק", " צ טנק   ")
    replace_header("חותמת זמן", "חותמת זמן.")
    replace_header("דוח זיווד [חבל פריסה]", "דוח זיווד [ חבל פריסה ]")
    ws.cell(row=1, column=ws.max_column + 1).value = "עמודה חדשה !"

    dest = tmp_path / "fuzzy.xlsx"
    wb.save(dest)

    responses, snapshot = FormResponsesAdapter.load_with_schema(dest)
    assert responses, "Expected responses parsed with normalized headers"
    assert any(m.family == "zivud" and "חבל פריסה" in m.item for m in snapshot.mapped)
    assert "עמודה חדשה !" in snapshot.unmapped
