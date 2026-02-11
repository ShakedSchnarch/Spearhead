from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from spearhead.v1.matrix_ingest import parse_matrix_workbook


def _write_minimal_matrix(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "שבוע 7"

    ws.cell(row=2, column=1, value="הפריט")
    ws.cell(row=2, column=2, value="תקן")
    ws.cell(row=2, column=3, value="צ'636")
    ws.cell(row=2, column=4, value="צ'653")
    ws.cell(row=2, column=19, value="תחמושת")

    ws.cell(row=3, column=1, value="חבל פריסה")
    ws.cell(row=3, column=2, value=1)
    ws.cell(row=3, column=3, value="קיים")
    ws.cell(row=3, column=4, value="בלאי")

    ws.cell(row=4, column=1, value="מקמש")
    ws.cell(row=4, column=2, value=1)
    ws.cell(row=4, column=3, value="קיים")
    ws.cell(row=4, column=4, value="חוסר")

    ws.cell(row=3, column=20, value="תקן")
    ws.cell(row=3, column=21, value="צ'636")
    ws.cell(row=3, column=22, value="צ'653")
    ws.cell(row=4, column=19, value="מאג")
    ws.cell(row=4, column=20, value=40)
    ws.cell(row=4, column=21, value=39)
    ws.cell(row=4, column=22, value=40)

    wb.save(path)


def test_parse_matrix_workbook_builds_per_tank_events(tmp_path: Path):
    workbook_path = tmp_path / "kfir.xlsx"
    _write_minimal_matrix(workbook_path)

    result = parse_matrix_workbook(
        workbook_path,
        company="Kfir",
        year=2026,
        source_prefix="kfir-sheet",
    )

    assert result.warnings == []
    assert result.sheets_processed == ["שבוע 7"]
    assert len(result.events) == 2

    payloads = {event["payload"]["צ טנק"]: event["payload"] for event in result.events}
    assert payloads["צ'636"]["דוח זיווד [חבל פריסה]"] == "קיים"
    assert payloads["צ'653"]["דוח זיווד [חבל פריסה]"] == "בלאי"
    assert payloads["צ'653"]["סטטוס ציוד קשר [מקמש]"] == "חוסר"
    assert payloads["צ'636"]["ברוסי מאג"] == "חוסר"
    assert payloads["צ'653"]["ברוסי מאג"] == "קיים"


def test_parse_matrix_workbook_ignores_non_week_sheets(tmp_path: Path):
    workbook_path = tmp_path / "no-weeks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "פורמט"
    ws.cell(row=1, column=1, value="No week data")
    wb.save(workbook_path)

    result = parse_matrix_workbook(workbook_path, company="Kfir", year=2026, source_prefix="kfir-sheet")
    assert result.events == []
    assert result.sheets_processed == []
