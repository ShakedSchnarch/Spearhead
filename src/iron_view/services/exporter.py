from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from iron_view.config import settings
from iron_view.services.analytics import FormAnalytics


class ExcelExporter:
    """
    Builds platoon and battalion Excel snapshots from stored form responses.
    The shape is intentionally simple: gaps for zivud, ammo totals with per-tank averages,
    means/comm gaps, and open issues with tank identifiers.
    """

    def __init__(self, analytics: FormAnalytics, output_dir: Optional[Path] = None):
        self.analytics = analytics
        self.output_dir = Path(output_dir or settings.paths.output_dir)

    def export_platoon(self, platoon: str, week: Optional[str] = None) -> Path:
        week_label = week or self.analytics.latest_week()
        if not week_label:
            raise ValueError("No week data found to export.")

        summary = self.analytics.summarize(week=week_label)["platoons"].get(platoon)
        if not summary:
            raise ValueError(f"No data found for platoon '{platoon}' in week '{week_label}'.")

        wb = Workbook()
        ws_zivud = wb.active
        ws_zivud.title = "זיווד"
        ws_zivud.append(["פריט", "חוסרים/בלאי"])
        if summary.zivud_gaps:
            for item, count in sorted(summary.zivud_gaps.items(), key=lambda kv: kv[1], reverse=True):
                ws_zivud.append([item, count])
        else:
            ws_zivud.append(["אין נתונים", None])
        _autosize(ws_zivud)

        ws_ammo = wb.create_sheet("תחמושת")
        ws_ammo.append(["אמצעי", "סה\"כ", "ממוצע לטנק"])
        if summary.ammo:
            for item, metrics in sorted(summary.ammo.items()):
                ws_ammo.append([item, metrics["total"], metrics["avg_per_tank"]])
        else:
            ws_ammo.append(["אין נתונים", None, None])
        _autosize(ws_ammo)

        ws_means = wb.create_sheet("אמצעים")
        ws_means.append(["אמצעי", "חוסרים/בלאי", "ממוצע לטנק"])
        if summary.means:
            for item, metrics in sorted(summary.means.items(), key=lambda kv: kv[1]["count"], reverse=True):
                ws_means.append([item, metrics["count"], metrics["avg_per_tank"]])
        else:
            ws_means.append(["אין נתונים", None, None])
        _autosize(ws_means)

        ws_issues = wb.create_sheet("פערי צלמים")
        ws_issues.append(["פריט", "צ טנק", "שם המט\"ק", "דגשים"])
        if summary.issues:
            for issue in summary.issues:
                ws_issues.append([issue["item"], issue["tank_id"], issue.get("commander"), issue["detail"]])
        else:
            ws_issues.append(["אין פערים", "", "", ""])
        _autosize(ws_issues)

        self.output_dir.mkdir(parents=True, exist_ok=True)
        path = self.output_dir / f"platoon_{platoon}_{week_label}.xlsx"
        wb.save(path)
        return path

    def export_battalion(self, week: Optional[str] = None) -> Path:
        week_label = week or self.analytics.latest_week()
        if not week_label:
            raise ValueError("No week data found to export.")

        summary = self.analytics.summarize(week=week_label)
        platoon_names = sorted(summary["platoons"].keys())
        wb = Workbook()

        # Zivud gaps per platoon + battalion total
        ws_zivud = wb.active
        ws_zivud.title = "זיווד גדודי"
        zivud_items = set()
        for p in summary["platoons"].values():
            zivud_items.update(p.zivud_gaps.keys())
        header = ["פריט"] + platoon_names + ["סה\"כ גדודי"]
        ws_zivud.append(header)
        if zivud_items:
            for item in sorted(zivud_items):
                row = [item]
                total = 0
                for platoon in platoon_names:
                    val = summary["platoons"][platoon].zivud_gaps.get(item, 0)
                    row.append(val)
                    total += val
                row.append(total)
                ws_zivud.append(row)
            ws_zivud.append(
                ["סה\"כ חוסרים/בלאי"]
                + [summary["battalion"]["zivud_gaps"].get(p, 0) for p in platoon_names]
                + [sum(summary["battalion"]["zivud_gaps"].values())]
            )
        else:
            ws_zivud.append(["אין נתונים"] + [""] * (len(platoon_names) + 1))
        _autosize(ws_zivud)

        # Ammo totals and per-tank averages per platoon + battalion
        ws_ammo = wb.create_sheet("תחמושת גדודית")
        ammo_items = set()
        for p in summary["platoons"].values():
            ammo_items.update(p.ammo.keys())
        header = ["אמצעי"]
        for platoon in platoon_names:
            header.extend([f"סה\"כ {platoon}", f"ממוצע לטנק {platoon}"])
        header.extend(["סה\"כ גדודי", "ממוצע לטנק גדודי"])
        ws_ammo.append(header)
        battalion_ammo = summary["battalion"]["ammo"]
        battalion_tanks = summary["battalion"]["tank_count"]
        if ammo_items:
            for item in sorted(ammo_items):
                row = [item]
                for platoon in platoon_names:
                    metrics = summary["platoons"][platoon].ammo.get(item, {"total": 0.0, "avg_per_tank": None})
                    row.extend([metrics.get("total", 0.0), metrics.get("avg_per_tank")])
                total = battalion_ammo.get(item, {}).get("total", 0.0)
                avg = battalion_ammo.get(item, {}).get("avg_per_tank") or (total / battalion_tanks if battalion_tanks else None)
                row.extend([total, avg])
                ws_ammo.append(row)
        else:
            ws_ammo.append(["אין נתונים"] + [""] * (len(platoon_names) * 2 + 2))
        _autosize(ws_ammo)

        # Means / comm gaps per platoon
        ws_means = wb.create_sheet("אמצעים")
        means_items = set()
        for p in summary["platoons"].values():
            means_items.update(p.means.keys())
        header = ["אמצעי"]
        for platoon in platoon_names:
            header.extend([f"חוסרים/בלאי {platoon}", f"ממוצע לטנק {platoon}"])
        header.extend(["סה\"כ גדודי", "ממוצע לטנק גדודי"])
        ws_means.append(header)
        battalion_means = summary["battalion"]["means"]
        battalion_tanks = summary["battalion"]["tank_count"]
        if means_items:
            for item in sorted(means_items):
                row = [item]
                for platoon in platoon_names:
                    metrics = summary["platoons"][platoon].means.get(item, {"count": 0, "avg_per_tank": None})
                    row.extend([metrics.get("count", 0), metrics.get("avg_per_tank")])
                total = battalion_means.get(item, {}).get("count", 0)
                avg = battalion_means.get(item, {}).get("avg_per_tank") or (total / battalion_tanks if battalion_tanks else None)
                row.extend([total, avg])
                ws_means.append(row)
        else:
            ws_means.append(["אין נתונים"] + [""] * (len(platoon_names) * 2 + 2))
        _autosize(ws_means)

        # Tank counts
        ws_tanks = wb.create_sheet("צי טנקים")
        ws_tanks.append(["פלוגה", "מספר טנקים"])
        for platoon in platoon_names:
            ws_tanks.append([platoon, summary["platoons"][platoon].tank_count])
        ws_tanks.append(["סה\"כ גדודי", summary["battalion"]["tank_count"]])
        _autosize(ws_tanks)

        self.output_dir.mkdir(parents=True, exist_ok=True)
        path = self.output_dir / f"battalion_{week_label}.xlsx"
        wb.save(path)
        return path

    def export_all_for_week(self, week: Optional[str] = None) -> dict:
        week_label = week or self.analytics.latest_week()
        if not week_label:
            raise ValueError("No week data found to export.")

        paths = {}
        for platoon in self.analytics.platoons(week=week_label):
            paths[f"platoon:{platoon}"] = self.export_platoon(platoon, week=week_label)
        paths["battalion"] = self.export_battalion(week=week_label)
        return paths


def _autosize(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 60)
