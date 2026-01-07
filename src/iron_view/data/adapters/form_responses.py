from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Any, Dict, Optional

from openpyxl import load_workbook

from iron_view.data.dto import FormResponseRow


class FormResponsesAdapter:
    """
    Parses the Google Form responses export (טופס דוחות סמפ כפיר. (תגובות)).
    The first row is headers; subsequent rows are responses with a mix of status strings and numbers.
    """

    @classmethod
    def load(cls, file_path: Path) -> List[FormResponseRow]:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [cls._safe_str(h) for h in rows[0]]
        records: List[FormResponseRow] = []

        for idx, row in enumerate(rows[1:], start=2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_dict: Dict[str, Any] = {}
            for h, v in zip(headers, row):
                if h:
                    row_dict[h] = v

            tank_id = cls._safe_str(row_dict.get("צ טנק"))
            timestamp = cls._parse_timestamp(row_dict.get("חותמת זמן") or row_dict.get("תאריך"))

            records.append(
                FormResponseRow(
                    source_file=file_path,
                    row_index=idx,
                    tank_id=tank_id,
                    timestamp=timestamp,
                    fields=row_dict,
                )
            )

        return records

    @staticmethod
    def _safe_str(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _parse_timestamp(value: Any) -> Optional[datetime]:
        if value is None:
            return None
        # openpyxl may parse Excel date numbers into datetime already
        if isinstance(value, datetime):
            return value
        try:
            # Handle Excel serial date as float/int
            if isinstance(value, (int, float)):
                # Excel serial date (from 1899-12-30)
                base = datetime(1899, 12, 30)
                return base + timedelta(days=float(value))
            return datetime.fromisoformat(str(value))
        except Exception:
            return None
