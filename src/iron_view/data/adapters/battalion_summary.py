from pathlib import Path
from typing import List, Optional, Sequence, Any

from openpyxl import load_workbook

from iron_view.data.dto import TabularRecord


class BattalionSummaryAdapter:
    """
    Parses the battalion-level summary sheet (מסמך דוחות גדודי).
    Focus on two sections: דוח זיווד (left) and תחמושת (right).
    Row 1 marks section starts; row 2 contains column headers; data starts at row 3.
    """

    SECTION_ZIVUD = "summary_zivud"
    SECTION_AMMO = "summary_ammo"

    @classmethod
    def load(cls, file_path: Path) -> List[TabularRecord]:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return []

        section_row = rows[0]
        header_row = rows[1]
        ammo_start = cls._find_section_start(section_row, "תחמושת")

        zivud_headers = header_row[:ammo_start] if ammo_start else header_row
        ammo_headers: Sequence[Any] = header_row[ammo_start:] if ammo_start is not None else []

        records: List[TabularRecord] = []

        for idx, row in enumerate(rows[2:], start=3):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            item = cls._safe_str(row[0]) if len(row) > 0 else None
            if item:
                records.extend(
                    cls._collect_row(
                        file_path=file_path,
                        section=cls.SECTION_ZIVUD,
                        row_index=idx,
                        item=item,
                        headers=zivud_headers,
                        row=row,
                        offset=0,
                    )
                )

            if ammo_headers:
                ammo_item = cls._safe_str(row[ammo_start]) if ammo_start is not None and ammo_start < len(row) else None
                if ammo_item:
                    records.extend(
                        cls._collect_row(
                            file_path=file_path,
                            section=cls.SECTION_AMMO,
                            row_index=idx,
                            item=ammo_item,
                            headers=ammo_headers,
                            row=row,
                            offset=ammo_start,
                        )
                    )

        return records

    @staticmethod
    def _find_section_start(row: Sequence[Any], label: str) -> Optional[int]:
        for i, v in enumerate(row):
            if v and str(v).strip() == label:
                return i
        return None

    @staticmethod
    def _safe_str(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @classmethod
    def _collect_row(
        cls,
        file_path: Path,
        section: str,
        row_index: int,
        item: str,
        headers: Sequence[Any],
        row: Sequence[Any],
        offset: int,
    ) -> List[TabularRecord]:
        records: List[TabularRecord] = []
        for j, header in enumerate(headers):
            header_text = cls._safe_str(header)
            if not header_text:
                continue
            if header_text == "הפריט":
                continue
            col_idx = offset + j
            value = row[col_idx] if col_idx < len(row) else None
            if value in (None, ""):
                continue
            records.append(
                TabularRecord(
                    source_file=file_path,
                    section=section,
                    item=item,
                    column=header_text,
                    value=value,
                    row_index=row_index,
                )
            )
        return records
