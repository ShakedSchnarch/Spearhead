import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Any, Dict, Optional, Tuple

from openpyxl import load_workbook

from spearhead.data.dto import FormResponseRow
from spearhead.data.field_mapper import FieldMapper, SchemaSnapshot
from spearhead.exceptions import DataSourceError

logger = logging.getLogger(__name__)


class FormResponsesAdapter:
    """
    Parses the Google Form responses export (טופס דוחות סמפ כפיר. (תגובות)).
    The first row is headers; subsequent rows are responses with a mix of status strings and numbers.
    Driven by config/fields.yaml to allow header drift and alias matching.
    """

    @classmethod
    def load(cls, file_path: Path, mapper: Optional[FieldMapper] = None, source_id: Optional[str] = None) -> List[FormResponseRow]:
        records, _ = cls.load_with_schema(file_path, mapper=mapper, source_id=source_id)
        return records

    @classmethod
    def load_with_schema(
        cls,
        file_path: Path,
        mapper: Optional[FieldMapper] = None,
        source_id: Optional[str] = None,
        platoon: Optional[str] = None,
    ) -> Tuple[List[FormResponseRow], SchemaSnapshot]:
        mapper = mapper or FieldMapper()
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            empty_snapshot = mapper.snapshot([])
            raise DataSourceError(cls._format_missing_required(mapper, ["tank_id", "timestamp"]))

        headers = [cls._safe_str(h) for h in rows[0]]
        snapshot = mapper.snapshot(headers)

        if snapshot.missing_required:
            raise DataSourceError(cls._format_missing_required(mapper, snapshot.missing_required))

        if snapshot.unmapped:
            logger.warning(
                "unmapped headers detected",
                extra={"headers": snapshot.unmapped, "source_file": str(file_path)},
            )

        if not platoon:
            platoon = mapper.infer_platoon(file_path, source_id=source_id)
        logger.info(f"DEBUG: Processing file {file_path.name}")
        logger.info(f"DEBUG: Inferred Platoon: '{platoon}' (Source ID: {source_id})")
        logger.info(f"DEBUG: Mapped Headers: {[m.to_dict() for m in snapshot.mapped]}")

        records: List[FormResponseRow] = []

        for idx, row in enumerate(rows[1:], start=2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_dict: Dict[str, Any] = {}
            for h, v in zip(headers, row):
                if h:
                    row_dict[h] = v

            tank_id = mapper.extract_tank_id(row_dict)
            timestamp_val = mapper.extract_by_aliases(row_dict, mapper.config.form.timestamp.aliases)
            timestamp = cls._parse_timestamp(timestamp_val)
            week_label = cls._week_label(timestamp)

            if idx <= 6: # Debug first 5 rows
                logger.info(f"DEBUG Row {idx}: TankID='{tank_id}', TS_Val='{timestamp_val}', TS_Parsed='{timestamp}', Week='{week_label}'")

            records.append(
                FormResponseRow(
                    source_file=file_path,
                    platoon=platoon,
                    row_index=idx,
                    tank_id=tank_id,
                    timestamp=timestamp,
                    week_label=week_label,
                    fields=row_dict,
                )
            )

        logger.info(f"DEBUG: Parsed {len(records)} records from {file_path.name}")
        return records, snapshot

    @staticmethod
    def _safe_str(value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _parse_timestamp(value: Any) -> Optional[datetime]:
        if value is None:
            return None
        # openpyxl may parse Excel date numbers into datetime already
        if isinstance(value, datetime):
            return value
        try:
            # Handle Excel serial date as float/int
            if isinstance(value, (int, float)):
                # Excel serial date (from 1899-12-30)
                base = datetime(1899, 12, 30)
                return base + timedelta(days=float(value))
            return datetime.fromisoformat(str(value))
        except Exception:
            return None

    @staticmethod
    def _week_label(ts: Optional[datetime]) -> Optional[str]:
        if not ts:
            return None
        try:
            return ts.strftime("%G-W%V")
        except Exception:
            return None

    @staticmethod
    def _format_missing_required(mapper: FieldMapper, missing: List[str]) -> str:
        missing_labels = {
            "tank_id": ", ".join(mapper.config.form.tank_id.aliases),
            "timestamp": ", ".join(mapper.config.form.timestamp.aliases),
        }
        detail_parts = []
        for key in missing:
            if key in missing_labels:
                detail_parts.append(f"{key} (expected aliases: {missing_labels[key]})")
            else:
                detail_parts.append(key)
        return f"Missing required columns: {', '.join(detail_parts)}"
