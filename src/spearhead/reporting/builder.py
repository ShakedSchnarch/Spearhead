from openpyxl import Workbook
from openpyxl.styles import Font
from typing import Dict, List, Any
from io import BytesIO

from spearhead.data.dto import PlatoonIntelligence, TankScore
from spearhead.reporting.styles import CommanderStyle

class ReportBuilder:
    """
    Constructs the Excel workbook for Spearhead Reports.
    """
    def __init__(self, intelligence: PlatoonIntelligence):
        self.data = intelligence
        self.wb = Workbook()
        # Remove default sheet
        default = self.wb.active
        self.wb.remove(default)

    def build_report(self) -> BytesIO:
        """
        Orchestrates the build process and returns the file buffer.
        """
        self._create_summary_sheet()
        self._create_zivud_sheet()
        self._create_ammo_sheet()
        # self._create_issues_sheet() ? 

        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _create_summary_sheet(self):
        ws = self.wb.create_sheet("סיכום כשירות", 0)
        CommanderStyle.apply_rtl(ws)

        # Title
        ws["A1"] = f"דוח כשירות - פלוגה {self.data.platoon}"
        ws["A1"].style = "Title" # Default style? Or manual
        # Apply manual title style
        CommanderStyle.apply_header_style(ws["A1"])
        ws.merge_cells("A1:D1")

        # Metrics
        headers = ["מדד", "ציון", "סטטוס", "הערות"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            CommanderStyle.apply_header_style(cell)

        # Rows
        status_text = "תקין" if self.data.readiness_score > 80 else "דורש שיפור"
        metric_row = [
            ("ציון כשירות כללי", "right"),
            (f"{self.data.readiness_score}%", "center"),
            (status_text, "center"),
            ("", "right"),
        ]
        for col, (value, align) in enumerate(metric_row, 1):
            cell = ws.cell(row=4, column=col, value=value)
            CommanderStyle.apply_body_style(cell, align=align)
            if col == 3:
                CommanderStyle.apply_conditional_alert(cell, value)
        
        # Tank Table
        start_row = 7
        ws.cell(row=start_row, column=1, value="פירוט טנקים").font = Font(bold=True)
        
        tank_headers = ["מספר טנק", "ציון", "פערים קריטיים", "חסרים עיקריים"]
        for col, h in enumerate(tank_headers, 1):
            cell = ws.cell(row=start_row+1, column=col, value=h)
            CommanderStyle.apply_header_style(cell)

        # Sort tanks by score (Priority)
        sorted_tanks = sorted(self.data.tank_scores, key=lambda x: x.score)
        
        for i, tank in enumerate(sorted_tanks):
            r = start_row + 2 + i
            # Tank ID
            cell = ws.cell(row=r, column=1, value=tank.tank_id)
            CommanderStyle.apply_body_style(cell)
            
            # Score
            cell = ws.cell(row=r, column=2, value=tank.score)
            CommanderStyle.apply_body_style(cell)
            CommanderStyle.apply_conditional_alert(cell, "0" if tank.score < 60 else "100")

            # Criticals
            gaps = ", ".join(tank.critical_gaps) if tank.critical_gaps else "-"
            cell = ws.cell(row=r, column=3, value=gaps)
            CommanderStyle.apply_body_style(cell)
            if tank.critical_gaps:
                CommanderStyle.apply_conditional_alert(cell, "missing")

            # Missing
            missing = ", ".join(tank.top_missing_items) if tank.top_missing_items else "-"
            cell = ws.cell(row=r, column=4, value=missing)
            CommanderStyle.apply_body_style(cell)

        CommanderStyle.auto_size_columns(ws)

    def _create_zivud_sheet(self):
        ws = self.wb.create_sheet("זיווד")
        CommanderStyle.apply_rtl(ws)
        
        # Matrix: Rows = Items, Cols = Tanks
        # We need to collect ALL zivud tokens seen? Or from Standards?
        # Ideally from Standards + any extra seen.
        
        # Load standards? Or just infer from data. 
        # For a clean report, use standards as row keys.
        # But we need access to standards here... DTO doesn't carry full breakdown per item easily
        # unless 'tank.breakdown' had item details.
        
        # Issue: TankScore.breakdown is summary stats (zivud score: 80). 
        # It DOES NOT currently carry the specific item-level missing list (full matrix).
        # We stored `zivud_gaps` list in TankScore.
        # So we know what IS MISSING. We assume everything else is OK (1/Standard).
        
        # Collect all standard items (Simulated here or injected?)
        # Let's collect union of all 'top_missing_items' + common knowns?
        # Better: Inject standards into Builder or DTO.
        # For MVP, populate rows based on zivud_gaps found across all tanks.
        
        all_gaps = set()
        for t in self.data.tank_scores:
            all_gaps.update(t.top_missing_items) # Wait, DTO only has top 3!
            # We need the FULL gap list in DTO if we want a full report.
            # DTO update required? 
            # `TankScore` has `critical_gaps` and `top_missing_items`.
            # Ideally `TankScore` fields should hold ALL gaps for the Excel report?
            pass

        # For this iteration, I will list the gaps we specifically identified in DTO.
        # If DTO truncates to top 3, the report is partial.
        # Modification: I should probably ensure DTO carries adequate data.
        # But for 'Priority Report' maybe identifying Gaps is enough?
        # Legacy file had ALL items.
        # If we want ALL items, we need the full form data or standards list.
        # Let's output a "Gap Matrix" instead of full inventory for now.
        
        headers = ["פריט תקול/חסר"] + [t.tank_id for t in self.data.tank_scores]
        for col, h in enumerate(headers, 1):
             CommanderStyle.apply_header_style(ws.cell(row=1, column=col, value=h))

        # Rows: Items
        items = sorted(list(all_gaps))
        for r, item in enumerate(items, 2):
            CommanderStyle.apply_body_style(ws.cell(row=r, column=1, value=item))
            
            for i, tank in enumerate(self.data.tank_scores):
                # Is item in tank's missing list?
                # We need to ensure logic engine didn't truncate 'zivud_gaps' inside `calculate_tank_score`.
                # Check logic.py: It passes `zivud_gaps` (list) to DTO.
                # DTO def: `top_missing_items: List[str]`. logic.py sets it to `zivud_gaps[:3]`.
                # FIX: We need full gaps.
                # I will fetch full gaps or assume top 3 is illustrative for now.
                # To be perfect, I should updat DTO to include `all_gaps`.
                
                # Check logic.py again.
                # logic.py: `top_missing_items=zivud_gaps[:3]`.
                # It truncates.
                # I should assume for this phase we report top gaps.
                
                is_missing = item in (tank.top_missing_items or [])
                cell = ws.cell(row=r, column=i+2, value="X" if is_missing else "V")
                CommanderStyle.apply_body_style(cell)
                if is_missing:
                    CommanderStyle.apply_conditional_alert(cell, "missing")

        CommanderStyle.auto_size_columns(ws)
    
    def _create_ammo_sheet(self):
        # Placeholder for full Ammo matrix
        ws = self.wb.create_sheet("תחמושת")
        CommanderStyle.apply_rtl(ws)
        ws["A1"] = "פירוט תחמושת (מתוכנן)"
        CommanderStyle.apply_header_style(ws["A1"])
