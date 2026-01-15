from openpyxl.styles import Font, PatternFill, Border, Alignment, Side, NamedStyle
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

class CommanderStyle:
    """
    Defines the visual language for Spearhead's "Commander Reports".
    Targeting a professional, military/tactical look (Grid lines, RTL, Clear Headers).
    """
    
    # Colors
    HEADER_BG = "4F81BD" # Safe Blue
    HEADER_TEXT = "FFFFFF"
    ALERT_BG = "FFC7CE" # Light Red
    ALERT_TEXT = "9C0006" # Dark Red
    
    # Borders
    THIN_BORDER = Side(border_style="thin", color="000000")
    BORDER_ALL = Border(top=THIN_BORDER, left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER)

    @staticmethod
    def apply_rtl(ws: Worksheet):
        """Sets the worksheet view direction to Right-to-Left."""
        ws.sheet_view.rightToLeft = True

    @staticmethod
    def apply_header_style(cell):
        """Applies header styling: Blue BG, White Bold text, Borders."""
        cell.font = Font(bold=True, color=CommanderStyle.HEADER_TEXT, name="Arial", size=11)
        cell.fill = PatternFill(start_color=CommanderStyle.HEADER_BG, end_color=CommanderStyle.HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CommanderStyle.BORDER_ALL

    @staticmethod
    def apply_body_style(cell, align="right"):
        """Applies standard table body styling."""
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = CommanderStyle.BORDER_ALL

    @staticmethod
    def apply_conditional_alert(cell, value):
        """
        Applies red styling if the value indicates a problem.
        Problems: "Missing", "0" (when expected), "No", "Faulty", Numbers < Standard?
        """
        val_str = str(value).lower().strip()
        is_bad = False
        
        bad_keywords = ["חסר", "missing", "תקלה", "לא תקין", "0"]
        if any(bad in val_str for bad in bad_keywords):
            is_bad = True
            
        # Specific numeric check? (Caller might handle logic, style just handles result?)
        # Let's trust caller triggers.
        
        if is_bad:
            cell.fill = PatternFill(start_color=CommanderStyle.ALERT_BG, end_color=CommanderStyle.ALERT_BG, fill_type="solid")
            cell.font = Font(color=CommanderStyle.ALERT_TEXT)

    @staticmethod
    def auto_size_columns(ws: Worksheet):
        """Simple auto-size heuristic."""
        for col in ws.columns:
            cells = [cell for cell in col if not isinstance(cell, MergedCell)]
            if not cells:
                continue
            max_length = 0
            column = cells[0].column_letter # Get the column name
            for cell in cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap width
