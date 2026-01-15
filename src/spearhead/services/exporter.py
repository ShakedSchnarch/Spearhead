from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from spearhead.config import settings
from spearhead.services.analytics import FormAnalytics
from spearhead.services.intelligence import IntelligenceService
from spearhead.reporting.builder import ReportBuilder


class ExcelExporter:
    """
    Builds platoon and battalion Excel snapshots using the new ReportBuilder.
    """

    def __init__(self, analytics: FormAnalytics, intelligence: IntelligenceService, output_dir: Optional[Path] = None):
        self.analytics = analytics
        self.intelligence = intelligence
        self.output_dir = Path(output_dir or settings.paths.output_dir)

    def export_platoon(self, platoon: str, week: Optional[str] = None) -> Path:
        # 1. Fetch Intelligence Data
        intel = self.intelligence.get_platoon_intelligence(platoon, week=week)
        
        # 2. Build Report
        builder = ReportBuilder(intel)
        buffer = builder.build_report()
        
        # 3. Save
        self.output_dir.mkdir(parents=True, exist_ok=True)
        safe_platoon = platoon.replace(" ", "_").replace("/", "-")
        week_label = intel.week.replace(" ", "_")
        filename = f"Spearhead_Commander_Report_{safe_platoon}_{week_label}.xlsx"
        path = self.output_dir / filename
        
        with open(path, "wb") as f:
            f.write(buffer.getvalue())
            
        return path

    def export_battalion(self, week: Optional[str] = None) -> Path:
        # Legacy fallback or future implementation
        # For now, keep simpler logic or raise NotImplemented if Builder not ready for Battalion?
        # Let's keep a minimal legacy-like implementation if possible or just rely on analytics for now?
        # User asked for "Files like the old ones". 
        # The Battalion file was aggregated. 
        # Plan: Refactor Battalion export later. Focus on Platoon (Tactical) first.
        # But we need basic functionality to pass tests/API calls.
        # I'll preserve the logic but wrap it cleanly or mark TODO.
        # Let's preserve the old logic for Battalion for safety, but using `self.analytics` which we kept.
        
        week_label = week or self.analytics.latest_week()
        if not week_label:
            raise ValueError("No week data found to export.")

        summary = self.analytics.summarize(week=week_label)
        platoon_names = sorted(summary["platoons"].keys())
        wb = Workbook()

        # Zivud gaps per platoon + battalion total
        ws_zivud = wb.active
        ws_zivud.title = "זיווד גדודי"
        ws_zivud.sheet_view.rightToLeft = True
        zivud_items = set()
        for p in summary["platoons"].values():
            zivud_items.update(p.zivud_gaps.keys())
        header = ["פריט"] + platoon_names + ["סה\"כ גדודי"]
        ws_zivud.append(header)
        for item in sorted(zivud_items):
            row = [item]
            total = 0
            for platoon in platoon_names:
                val = summary["platoons"][platoon].zivud_gaps.get(item, 0)
                row.append(val)
                total += val
            row.append(total)
            ws_zivud.append(row)
        _autosize(ws_zivud)
        
        self.output_dir.mkdir(parents=True, exist_ok=True)
        filename = f"Spearhead_Export_Battalion_{week_label}.xlsx"
        path = self.output_dir / filename
        wb.save(path)
        return path

    def export_all_for_week(self, week: Optional[str] = None) -> dict:
        week_label = week or self.analytics.latest_week()
        paths = {}
        # Get list from repo via analytics
        # Note: IntelligenceService handles single platoons well.
        for platoon in self.analytics.platoons(week=week_label):
            paths[f"platoon:{platoon}"] = self.export_platoon(platoon, week=week_label)
        paths["battalion"] = self.export_battalion(week=week_label)
        return paths


def _autosize(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 60)
