from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

WEEK_SHEET_RE = re.compile(r"שבוע\s*(\d+)")
TANK_ID_RE = re.compile(r"(\d{3,4})")

COMMUNICATION_ITEMS = {
    "מקמש",
    "גנטקס",
    "פתיל",
    "מעד",
    "מיק חירום",
    "רמק",
    "מבן",
    "אנטנת מבן",
    "NFC",
    "מכלול",
    "נר לילה",
    "קפ שליטה",
    "משיב מיקום",
    "מקמש\\מגן מכלול",
    "מדיה\\נר לילה",
}

AMMO_FIELD_ALIASES = {
    "מאג": "ברוסי מאג",
    "0.5": "ברוסי 05",
    "רימון רסס": "רימוני רסס",
    "רימון עשן": "רימוני עשן",
    "שאקל 25": "שאקל 25 טון",
    "שאקל 5": "שאקל 5 טון",
    'סט שיני חזיר\\נג"ח': 'שיני חזיר\\נג"ח',
}

IGNORED_ITEMS = {
    "הפריט",
    "תקן",
    "תחמושת",
    "דוח זיווד",
    "דוח זיווד ",
    "דוח ציוד קשר- יש לסמן בלאי\\ חוסר בצבע",
    "דוח צלם- נוספים",
    "ציוד משרד",
    "ציוד קשפל",
    "ציוד רנגלר",
    "הצלם",
    "צ'",
    "תקין",
    "תקין\\פירוט התקלה",
    "מדוחות",
    "מדוכות",
}


@dataclass
class MatrixIngestParseResult:
    events: list[dict[str, Any]]
    sheets_processed: list[str]
    warnings: list[str]


def parse_matrix_workbook(
    workbook_path: Path,
    *,
    company: str,
    year: int | None = None,
    source_prefix: str | None = None,
) -> MatrixIngestParseResult:
    wb = load_workbook(workbook_path, data_only=True)
    target_year = year or datetime.now(UTC).year
    source_root = source_prefix or workbook_path.stem

    events: list[dict[str, Any]] = []
    sheets_processed: list[str] = []
    warnings: list[str] = []

    for ws in wb.worksheets:
        match = WEEK_SHEET_RE.search(ws.title or "")
        if not match:
            continue

        week_num = int(match.group(1))
        sheet_events = _parse_week_sheet(
            ws,
            week_num=week_num,
            year=target_year,
            company=company,
            source_prefix=source_root,
        )
        if not sheet_events:
            warnings.append(f"No tank events parsed from sheet '{ws.title}'")
            continue
        events.extend(sheet_events)
        sheets_processed.append(ws.title)

    events.sort(key=lambda item: str(item.get("source_id", "")))
    return MatrixIngestParseResult(events=events, sheets_processed=sheets_processed, warnings=warnings)


def _parse_week_sheet(
    ws: Worksheet,
    *,
    week_num: int,
    year: int,
    company: str,
    source_prefix: str,
) -> list[dict[str, Any]]:
    timestamp = _week_timestamp(year=year, week_num=week_num)
    left_tanks = _collect_tank_columns(ws, row=2, min_col=3, max_col=18)
    right_tanks = _collect_tank_columns(ws, row=3, min_col=20, max_col=ws.max_column)
    all_tanks = sorted(set(left_tanks.values()) | set(right_tanks.values()))
    if not all_tanks:
        return []

    base_payload = {
        tank: {
            "צ טנק": tank,
            "חותמת זמן": timestamp.isoformat(),
            "פלוגה": company,
        }
        for tank in all_tanks
    }

    for row_idx in range(3, ws.max_row + 1):
        left_item = _clean_text(ws.cell(row=row_idx, column=1).value)
        if _is_valid_item(left_item):
            std_left = ws.cell(row=row_idx, column=2).value
            field_name = _left_field_name(left_item)
            for col_idx, tank in left_tanks.items():
                status = _normalize_status(ws.cell(row=row_idx, column=col_idx).value, std_left)
                if status is not None:
                    base_payload[tank][field_name] = status

        right_item = _clean_text(ws.cell(row=row_idx, column=19).value)
        if _is_valid_item(right_item):
            std_right = ws.cell(row=row_idx, column=20).value
            field_name = _right_field_name(right_item)
            for col_idx, tank in right_tanks.items():
                status = _normalize_status(ws.cell(row=row_idx, column=col_idx).value, std_right)
                if status is not None:
                    base_payload[tank][field_name] = status

    events: list[dict[str, Any]] = []
    for tank, payload in base_payload.items():
        # Keep only events with at least one operational field beyond metadata.
        if len(payload) <= 3:
            continue
        events.append(
            {
                "schema_version": "v2",
                "source_id": f"{source_prefix}:{ws.title}:{tank}",
                "received_at": timestamp.isoformat(),
                "payload": payload,
            }
        )
    return events


def _collect_tank_columns(
    ws: Worksheet,
    *,
    row: int,
    min_col: int,
    max_col: int,
) -> dict[int, str]:
    tanks: dict[int, str] = {}
    for col_idx in range(min_col, max_col + 1):
        cell_val = ws.cell(row=row, column=col_idx).value
        tank = _extract_tank(cell_val)
        if tank:
            tanks[col_idx] = tank
    return tanks


def _extract_tank(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    match = TANK_ID_RE.search(text)
    if not match:
        return None
    return f"צ'{match.group(1)}"


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _is_valid_item(item: str | None) -> bool:
    if not item:
        return False
    normalized = item.strip().lower()
    if not normalized:
        return False
    if normalized in {token.lower() for token in IGNORED_ITEMS}:
        return False
    if normalized.startswith("סה"):
        return False
    if "דוח" in normalized and "פלוגת" in normalized:
        return False
    if _extract_tank(item):
        return False
    if re.fullmatch(r"\d+\(.*\)", item):
        return False
    return True


def _left_field_name(item: str) -> str:
    if item in COMMUNICATION_ITEMS:
        return f"סטטוס ציוד קשר [{item}]"
    return f"דוח זיווד [{item}]"


def _right_field_name(item: str) -> str:
    if item in COMMUNICATION_ITEMS:
        return f"סטטוס ציוד קשר [{item}]"
    return AMMO_FIELD_ALIASES.get(item, item)


def _parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _numeric_to_status(value_num: float, standard_num: float | None) -> str:
    if standard_num is not None and standard_num > 0:
        return "קיים" if value_num >= standard_num else "חוסר"
    return "קיים" if value_num > 0 else "חוסר"


def _normalize_status(value: Any, standard: Any) -> str | None:
    if value is None:
        return None
    text = _clean_text(value)
    if text is None:
        return None

    lower = text.lower()
    if "לא קיים" in lower or "אין" in lower or "חוסר" in lower or "חסר" in lower:
        return "חוסר"
    if "בלאי" in lower or "תקול" in lower:
        return "בלאי"
    if "קיים" in lower or "תקין" in lower or lower == "יש":
        return "קיים"

    value_num = _parse_number(value)
    if value_num is not None:
        standard_num = _parse_number(standard)
        return _numeric_to_status(value_num, standard_num)

    return text


def _week_timestamp(*, year: int, week_num: int) -> datetime:
    # Week snapshots are treated as Monday 08:00 UTC for deterministic week labels.
    dt = datetime.fromisocalendar(year, week_num, 1)
    return dt.replace(hour=8, minute=0, second=0, microsecond=0, tzinfo=UTC)
